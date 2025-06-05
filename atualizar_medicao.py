from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import pandas as pd
import os

import sys

# Se argumentos forem passados via linha de comando, use-os:
if len(sys.argv) >= 3:
    caminho_pdf_dados = sys.argv[1]  # informacoes_extraidas.xlsx
    caminho_saida = sys.argv[2]      # medicao_atualizada.xlsx
    caminho_medicao = caminho_saida  # usado no script


# PRESERVAR A1 E A14 DA ABA RESUMO
valor_a1_original = ""
valor_a14_original = ""
try:
    if os.path.exists(caminho_medicao):
        wb_existente = load_workbook(caminho_medicao)
        if "RESUMO" in wb_existente.sheetnames:
            ws_resumo_antigo = wb_existente["RESUMO"]
            valor_a1_original = ws_resumo_antigo["A1"].value or ""
            valor_a14_original = ws_resumo_antigo["A14"].value or ""
except Exception as e:
    print("⚠️ Erro ao tentar preservar A1 e A14:", e)

if os.path.exists(caminho_medicao):
    df_base = pd.read_excel(caminho_medicao)
    df_base.columns = df_base.columns.str.strip()
    df_base["OSES"] = df_base["OSES"].astype(str).str.strip()
    df_base = df_base[df_base["OSES"].notna()]
    df_base = df_base[df_base.dropna(how="all").notna().any(axis=1)]
    df_base.reset_index(drop=True, inplace=True)
    print("📄 Planilha existente carregada.")
else:
    print("⚠️ Planilha de medição não encontrada. Criando uma nova...")
    df_base = pd.DataFrame(columns=[
        'OSES', 'COD SERVIÇO', 'ENDEREÇO', 'N°', 'BAIRRO',
        'DATA DE EXECUÇÃO', 'VALOR DA O.S',
        'MEDIDO/NÃO MEDIDO/PENDENCIA', 'ARQUIVO',
        'VALOR PAGO', 'DIFERENÇA DE VALOR',
        'Observações do Script', 'Valor PAGO Original'
    ])
def encontrar_coluna(colunas, nome_alvo):
    for col in colunas:
        if col.strip().lower() == nome_alvo.lower():
            return col
    return None

col_os_base = encontrar_coluna(df_base.columns, "OSES")
df_novos = pd.read_excel(caminho_pdf_dados)
df_novos.columns = df_novos.columns.str.strip()
col_os_pdf = encontrar_coluna(df_novos.columns, "Número da O.S.")
col_status_pdf = encontrar_coluna(df_novos.columns, "Arquivo")
col_status_base = encontrar_coluna(df_base.columns, "Arquivo")
col_valor_pdf = encontrar_coluna(df_novos.columns, "Valor da O.S.")
col_valor_base = encontrar_coluna(df_base.columns, "VALOR DA O.S")
col_valor_pago = encontrar_coluna(df_base.columns, "VALOR PAGO")
col_diferenca = encontrar_coluna(df_base.columns, "DIFERENÇA DE VALOR")

# Conversões de tipo reforçadas (agora no local certo)
df_base["OSES"] = df_base["OSES"].astype(str).str.strip()
df_novos[col_os_pdf] = df_novos[col_os_pdf].astype(str).str.strip()
df_base[col_valor_base] = pd.to_numeric(df_base[col_valor_base], errors="coerce")
df_base[col_valor_pago] = pd.to_numeric(df_base[col_valor_pago], errors="coerce")
df_novos[col_valor_pdf] = pd.to_numeric(df_novos[col_valor_pdf], errors="coerce")
col_medido = "MEDIDO/NÃO MEDIDO/PENDENCIA"
col_obs = "Observações do Script"
if col_obs in df_base.columns:
    df_base[col_obs] = df_base[col_obs].astype("string")
col_valor_original = "Valor PAGO Original"

for nome, valor in {
    "col_os_base": col_os_base,
    "col_os_pdf": col_os_pdf,
    "col_status_pdf": col_status_pdf,
    "col_status_base": col_status_base,
    "col_valor_pdf": col_valor_pdf,
    "col_valor_base": col_valor_base,
    "col_valor_pago": col_valor_pago,
    "col_diferenca": col_diferenca,
}.items():
    if not valor:
        raise ValueError(f"❌ Coluna obrigatória não encontrada: {nome}")

df_novos[col_os_pdf] = df_novos[col_os_pdf].astype(str).str.strip()

if col_medido not in df_base.columns:
    df_base[col_medido] = ""
if col_obs not in df_base.columns:
    df_base[col_obs] = ""
if col_valor_original not in df_base.columns:
    df_base[col_valor_original] = None

os_novos_set = set(df_novos[col_os_pdf].dropna().astype(str).str.strip())

for _, nova in df_novos.iterrows():
    if pd.isna(nova[col_os_pdf]):
        continue
    os_num = str(nova[col_os_pdf]).strip()
    valor_novo = nova[col_valor_pdf]
    status_novo = str(nova[col_status_pdf]).strip().upper()
    encontrados = df_base[df_base[col_os_base] == os_num]
    if not encontrados.empty:
        pos = encontrados.index[0]
        status_atual = str(df_base.loc[pos, col_status_base]).strip().upper()
        valor_pago = df_base.loc[pos, col_valor_pago]
        if status_atual == "PAGO":
            if status_novo != "PAGO":
                df_base.at[pos, col_status_base] = status_novo
                df_base.at[pos, col_medido] = "NÃO MEDIDO"
                df_base.at[pos, col_obs] = "Status alterado após pagamento"
                df_base.at[pos, col_valor_pago] = None
            elif valor_novo != valor_pago:
                df_base.at[pos, col_valor_original] = valor_pago
                df_base.at[pos, col_valor_pago] = valor_novo
                df_base.at[pos, col_obs] = "Valor pago atualizado após divergência"
        else:
            df_base.at[pos, col_status_base] = status_novo
            # Atualiza o VALOR DA O.S se o status NÃO for PAGO
            if status_novo != "PAGO":
                df_base.at[pos, col_valor_base] = valor_novo
            df_base.at[pos, col_valor_pago] = valor_novo if status_novo == "PAGO" else None
            df_base.at[pos, col_medido] = "MEDIDO" if status_novo == "PAGO" else "NÃO MEDIDO"
            df_base.at[pos, col_obs] = ""
    else:
        nova_linha = {
            col_os_base: os_num,
            "COD SERVIÇO": nova[encontrar_coluna(df_novos.columns, "Código do Serviço")],
            "ENDEREÇO": f"{nova['Rua']}, {nova['Número']}, {nova['Bairro']}",
            "N°": nova["Número"],
            "BAIRRO": nova["Bairro"],
            "DATA DE EXECUÇÃO": nova["Data de Baixa"],
            col_valor_base: valor_novo,
            col_valor_pago: valor_novo if status_novo == "PAGO" else None,
            col_status_base: status_novo,
            col_medido: "MEDIDO" if status_novo == "PAGO" else "NÃO MEDIDO",
            col_obs: "",
            col_valor_original: None
        }
        nova_df = pd.DataFrame([nova_linha])
        nova_df = nova_df.dropna(axis=1, how="all")  # Remove colunas totalmente vazias
        if not nova_df.dropna(how="all").empty:
            df_base = pd.concat([df_base, nova_df], ignore_index=True)

linhas_nao_encontradas_novas = []

# Limpa observações antigas se a O.S. voltou a aparecer
for idx, linha in df_base.iterrows():
    os_num = linha[col_os_base]
    if os_num in os_novos_set and str(linha[col_obs]).strip() == "Não encontrada nos novos relatórios":
        df_base.at[idx, col_obs] = ""

for idx, linha in df_base.iterrows():
    os_num = linha[col_os_base]
    if pd.notna(os_num) and os_num not in os_novos_set:
        if pd.isna(linha[col_obs]) or linha[col_obs] == "":
            df_base.at[idx, col_obs] = "Não encontrada nos novos relatórios"
            linhas_nao_encontradas_novas.append(idx)

# Atualização de status "EM FISCALIZAÇÃO" para "EM FISCALIZAÇÃO COM PENDENCIA" com base em rua, número e bairro
def normaliza_endereco_partes(row):
    rua = str(row.get("Rua") or row.get("ENDEREÇO")).split(",")[0].strip().upper()
    numero = str(row["N°"]).strip()
    bairro = str(row["BAIRRO"]).strip().upper()
    return f"{rua}, {numero}, {bairro}"

pendencias = df_base[df_base[col_status_base].str.upper().str.strip() == "EM PENDENCIA"]
fiscalizacao = df_base[df_base[col_status_base].str.upper().str.strip() == "EM FISCALIZAÇÃO"]

enderecos_pendencia = set(pendencias.apply(normaliza_endereco_partes, axis=1))

for idx_fisc, row_fisc in fiscalizacao.iterrows():
    endereco_fisc = normaliza_endereco_partes(row_fisc)
    if endereco_fisc in enderecos_pendencia:
        df_base.at[idx_fisc, col_status_base] = "EM FISCALIZAÇÃO COM PENDENCIA"

df_base[col_diferenca] = df_base.apply(
    lambda row: row[col_valor_pago] - row[col_valor_base]
    if pd.notna(row[col_valor_pago]) and pd.notna(row[col_valor_base]) and
       str(row[col_status_base]).strip().upper() == "PAGO" else "", axis=1
)

df_base.drop(columns=list(filter(lambda col: "RAZAO" in col.upper() or "UNNAMED" in col.upper(), df_base.columns)), inplace=True)
df_base = df_base[df_base[col_os_base].notna()].reset_index(drop=True)


# Ordenar conforme solicitado

# Adiciona colunas auxiliares para ordenação personalizada
df_base["__OBS_PRIORIDADE__"] = df_base["Observações do Script"].apply(
    lambda x: 0 if str(x).strip() == "Não encontrada nos novos relatórios" else (1 if pd.notna(x) and str(x).strip() != "" else 2)
)

# Ordenação personalizada incluindo prioridade de observações
df_base.sort_values(
    by=["MEDIDO/NÃO MEDIDO/PENDENCIA", "__OBS_PRIORIDADE__", "ARQUIVO", "VALOR DA O.S"],
    ascending=[True, True, True, False],
    inplace=True
)

# Remover a coluna auxiliar após a ordenação
df_base.drop(columns=["__OBS_PRIORIDADE__"], inplace=True)


df_base.to_excel(caminho_saida, index=False)

# FORMATAÇÃO
wb = load_workbook(caminho_saida)
ws_principal = wb.active
ws_principal.title = "Sheet1"
ws_principal.freeze_panes = "A2"  # Congela a primeira linha
ws_principal.auto_filter.ref = ws_principal.dimensions  # Adiciona filtro na primeira linha
# ZOOM DA PLANILHA
ws_principal.sheet_view.zoomScale = 70

# ALTERAÇÕES DE LARGURA SOLICITADAS
ws_principal.column_dimensions['A'].width = 14
ws_principal.column_dimensions['B'].width = 40
ws_principal.column_dimensions['C'].width = 55
ws_principal.column_dimensions['D'].width = 12
ws_principal.column_dimensions['E'].width = 38
ws_principal.column_dimensions['G'].width = 25
ws_principal.column_dimensions['D'].width = 5


colunas_largura = {
    'A': 20,
    'B': 33,
    'C': 45,
    'D': 12,
    'E': 32,
    'F': 17.11,
    'G': 18,
    'H': 18.33,
    'I': 21,
    'J': 13,
    'K': 11.89,
    'L': 38.00,
    'M': 13.56
}
for col, largura in colunas_largura.items():
    ws_principal.column_dimensions[col].width = largura
for i in range(1, ws_principal.max_row + 1):
    ws_principal.row_dimensions[i].height = 18


ws_principal.row_dimensions[1].height = 25.2
for col in range(1, 14):  # Colunas A (1) até M (13)
    cell = ws_principal.cell(row=1, column=col)
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")




fonte_normal = Font(name="Calibri", size=11)
fonte_negrito = Font(name="Calibri", size=11, bold=True)
borda_fina = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))


for i, row in enumerate(ws_principal.iter_rows(min_row=1, max_row=ws_principal.max_row, min_col=1, max_col=13), start=1):
    for cell in row:
        cell.font = fonte_negrito if i == 1 else fonte_normal
        cell.border = borda_fina
        if i == 1 and cell.value:
            cell.fill = PatternFill(start_color="FF8CC450", end_color="FF8CC450", fill_type="solid")

        cell.font = fonte_negrito if i == 1 else fonte_normal
        cell.border = borda_fina

formato_moeda = 'R$ #,##0.00'

# Aplicar formatação nas colunas específicas
for row in ws_principal.iter_rows(min_row=2, max_row=ws_principal.max_row):
    if isinstance(row[0].value, (int, float)):
        row[0].number_format = '0'
    for idx in [6, 9, 10]:
        if isinstance(row[idx].value, (int, float)):
            row[idx].number_format = formato_moeda

# RESUMO
ws_resumo = wb.create_sheet("RESUMO")
tabela = df_base.copy()
tabela[col_diferenca] = pd.to_numeric(tabela[col_diferenca], errors="coerce")
tabela[col_status_base] = tabela[col_status_base].astype(str).str.upper().str.strip()
is_corte = tabela[col_diferenca].apply(lambda x: isinstance(x, (int, float)) and x < 0)
resumo_base = tabela.groupby(col_status_base)[col_valor_base].sum().to_dict()
valor_enviado = tabela[col_valor_base].sum()
valor_cortes = abs(tabela[is_corte][col_diferenca].astype(float).sum())
valor_osnp = tabela[tabela[col_status_base] == "OSNP"][col_valor_base].sum()
valor_previsto = (
    tabela[tabela[col_status_base] == "PAGO"][col_valor_pago].dropna().sum() +
    tabela[tabela[col_status_base] == "EM CONFERENCIA"][col_valor_base].sum() +
    tabela[tabela[col_status_base] == "EM FISCALIZAÇÃO"][col_valor_base].sum() +
    tabela[tabela[col_status_base] == "EM DIVERGENCIA"][col_valor_base].sum()
)
valor_pago_total = tabela[col_valor_pago].dropna().sum()

ordem_status = [
    "ENVIADO", "PAGO", "EM CONFERENCIA", "EM FISCALIZAÇÃO", "EM DIVERGENCIA",
    "ABERTA",
    "EM PENDENCIA", "EM FISCALIZAÇÃO COM PENDENCIA", "OSNP", "CORTES",
    "PREVISAO DE MEDIÇÃO", 
]


resumo = []
for status in ordem_status:
    if status == "ENVIADO":
        valor = valor_enviado
    elif status == "PAGO":
        valor = valor_pago_total
    elif status == "CORTES":
        valor = valor_cortes
    elif status == "OSNP":
        valor = valor_osnp
    elif status == "PREVISAO DE MEDIÇÃO":
        valor = valor_previsto
    else:
        valor = resumo_base.get(status, 0)
    qtd = tabela[tabela[col_status_base].str.upper() == status.upper()].shape[0]
    resumo.append({"STATUS DA O.S": status, "QTD": qtd, "VALOR": valor})

# Atualiza a QTD da linha "CORTES" com a contagem real de O.S com diferença negativa
tabela[col_diferenca] = pd.to_numeric(tabela[col_diferenca], errors="coerce")
qtd_os_com_corte = tabela[tabela[col_diferenca] < 0].shape[0]
for i, row_data in enumerate(resumo):
    if row_data["STATUS DA O.S"].strip().upper() == "CORTES":
        resumo[i]["QTD"] = qtd_os_com_corte




# Preenche aba RESUMO
ws_resumo.append(["STATUS DA O.S", "QTD DE O.S", "VALOR"])
for col in ["A", "B", "C"]:
    ws_resumo[f"{col}1"].font = fonte_negrito

for row in resumo:
    ws_resumo.append([row["STATUS DA O.S"], row["QTD"], abs(row["VALOR"])])

ws_resumo.column_dimensions['A'].width = 30.89
ws_resumo.column_dimensions['B'].width = 12
ws_resumo.column_dimensions['C'].width = 14.22
ws_resumo.row_dimensions[1].height = 18.6


# Formatação da aba RESUMO
ws_resumo.merge_cells('A1:C1')
cell_title = ws_resumo['A1']
cell_title.value = valor_a1_original
cell_title.font = Font(name="Calibri", size=16, bold=True)
cell_title.alignment = Alignment(horizontal="center", vertical="center")
cell_title.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ws_resumo.row_dimensions[1].height = 25

# Cabeçalhos
ws_resumo['A2'].value = "STATUS DA O.S"
ws_resumo['B2'].value = "QTD DE O.S"
ws_resumo['C2'].value = "VALOR"
for col in ['A2', 'B2', 'C2']:
    cell = ws_resumo[col]
    cell.font = Font(name="Calibri", size=12, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = borda_fina
ws_resumo.row_dimensions[2].height = 20

# Dados
for i, row_data in enumerate(resumo, start=3):
    ws_resumo[f"A{i}"] = row_data["STATUS DA O.S"]
    ws_resumo[f"B{i}"] = row_data["QTD"]
    ws_resumo[f"C{i}"] = abs(row_data["VALOR"])
    for col in ['A', 'B', 'C']:
        cell = ws_resumo[f"{col}{i}"]
        cell.font = fonte_normal
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda_fina
        if col == 'C' and isinstance(cell.value, (int, float)):
            cell.number_format = 'R$ #,##0.00'

# Cores de preenchimento conforme status
status_colors = {
    "PAGO": "C6EFCE",
    "ENVIADO": "BDD7EE",
    "EM CONFERENCIA": "FFF2CC",
    "EM FISCALIZAÇÃO": "FFF2CC",
    "EM DIVERGENCIA": "FFF2CC",
    "ABERTA": "D9E1F2",
    "EM PENDENCIA": "F9CB9C",
    "EM FISCALIZAÇÃO COM PENDENCIA": "F9CB9C",
    "PREVISAO DE MEDIÇÃO": "D9EAD3",
    "CORTES": "F9CB9C",
    "OSNP": "F9CB9C"
}
for i, row_data in enumerate(resumo, start=3):
    status = row_data["STATUS DA O.S"].upper()
    cor = status_colors.get(status, None)
    if cor:
        for col in ['A', 'B', 'C']:
            ws_resumo[f"{col}{i}"].fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")

# Linha final de destaque (exemplo)
final_row = len(resumo) + 3
ws_resumo.merge_cells(f"A{final_row}:C{final_row}")
final_cell = ws_resumo[f"A{final_row}"]
final_cell.value = valor_a14_original
final_cell.font = Font(name="Calibri", size=12, bold=True)
final_cell.alignment = Alignment(horizontal="center", vertical="center")
final_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
final_cell.border = borda_fina

# Ajuste de colunas
ws_resumo.column_dimensions['A'].width = 35
ws_resumo.column_dimensions['B'].width = 14
ws_resumo.column_dimensions['C'].width = 20

fills = {
    "PAGO": "C6EFCE",
    "ENVIADO": "BDD7EE",
    "EM CONFERENCIA": "FFF2CC",
    "EM FISCALIZAÇÃO": "FFF2CC",
    "EM DIVERGENCIA": "FFF2CC",
    "ABERTA": "D9E1F2",
    "EM PENDENCIA": "F9CB9C",
    "EM FISCALIZAÇÃO COM PENDENCIA": "F9CB9C",
    "PREVISAO DE MEDIÇÃO": "D9EAD3",
    "CORTES": "F9CB9C",
    "OSNP": "F9CB9C"
}
for row in ws_resumo.iter_rows(min_row=2, max_row=ws_resumo.max_row):
    status = row[0].value
    cor = fills.get(status.upper(), None)
    if cor:
        for cell in row:
            if row[0].row not in [5, 6, 7, 8]:
                cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")

# Centralizar colunas A até M, exceto B (sem alterar outras formatações)
for row in ws_principal.iter_rows(min_row=2, max_row=ws_principal.max_row):
    for col_idx in range(1, 14):  # Colunas A (1) até M (13)
        if col_idx == 2:  # Ignora coluna B
            continue
        cell = row[col_idx - 1]
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=cell.alignment.wrap_text if cell.alignment else None
        )


borda_espessa = Border(
    left=Side(style="thick"),
    right=Side(style="thick"),
    top=Side(style="thick"),
    bottom=Side(style="thick")
)

# Determinar o intervalo da tabela de resumo
inicio_row = 1  # Início na linha 1 agora
fim_row = final_row  # Linha amarela final
inicio_col = 1  # Coluna A
fim_col = 3     # Coluna C

# Aplicar borda espessa ao redor da tabela

# Remover preenchimento das linhas 5 a 8 da aba RESUMO
for row in range(5, 9):
    for col in range(1, 4):
        cell = ws_resumo.cell(row=row, column=col)
        cell.fill = PatternFill(fill_type=None)
for row in range(inicio_row, fim_row + 1):
    for col in range(inicio_col, fim_col + 1):
        cell = ws_resumo.cell(row=row, column=col)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )


wb.save(caminho_saida)

# Garantir formatação correta da coluna "OSES" (coluna A)
# Garantir formatação correta da coluna "OSES" (coluna A)
try:
    wb_check = load_workbook(caminho_saida)
    ws_check = wb_check.active
    col_idx = 1  # Coluna A
    for row in ws_check.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        cell = row[0]
        if isinstance(cell.value, str) and cell.value.isdigit():
            cell.value = int(cell.value)
            cell.number_format = '0'
    wb_check.save(caminho_saida)

    
    
    from openpyxl.styles import PatternFill

    fill_vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Mapear todas as O.S. que devem ser destacadas
    oses_para_destacar = set(df_base.loc[linhas_nao_encontradas_novas, col_os_base])

    for row in ws_check.iter_rows(min_row=2, max_row=ws_check.max_row):
        os_valor = str(row[0].value).strip() if row[0].value is not None else ""
        if os_valor in oses_para_destacar:
            obs_cell = row[11]  # coluna "Observações do Script" (coluna L)
            obs_cell.fill = fill_vermelho

    wb_check.save(caminho_saida)


except Exception as e:
    print("⚠️ Erro ao aplicar formatação na coluna A:", e)